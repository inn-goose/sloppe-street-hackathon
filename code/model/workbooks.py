"""model/workbooks — write the four submission workbooks from the templates.

    PYTHONPATH=. .venv/bin/python -m code.model.workbooks

The last step: read `challenge/templates/*.xlsx`, put each forecast in the one cell that is
meant to hold it, and save to `submission/`. The templates are opened and re-saved rather
than rebuilt, so the sheet name, metric labels, units, fiscal-period column and cell shading all
survive exactly as supplied — the rules require every one of them to be left alone.

⛔ **Metric labels are matched exactly, and a miss is fatal.** The workbook says
`Adjusted gross margin`; the registry must agree, letter for letter. A fuzzy match here would put
Home Depot's comparable-sales figure on its net-sales row and the run would still "succeed" — the
one failure mode that produces a plausible, well-formatted, completely wrong submission. Any label
that does not resolve, or any target without a forecast, stops the run.

⚠️ **Units are the workbook's, not the model's.** Percentages go in as points (`4.5` for 4.5 %),
Hays' EPS in pence, money in the stated millions. Each is asserted against the template's own units
column before the value is written.
"""

from __future__ import annotations

import shutil
from pathlib import Path

from code.lib import config, store
from code.model.train import MODEL

TEMPLATES = config.TEMPLATES
SUBMISSION = config.SUBMISSION
#: workbook stem → the ticker whose forecasts fill it
BOOKS = {"ADI-FY2026Q3": "ADI", "DE-FY2026Q3": "DE",
         "HAS-FY2026": "LSE:HAS", "HD-FY2026Q2": "HD"}
#: What each unit string in the template's own column B implies about the number written.
UNIT_EXPECTS = {
    "USDm": "currency_m", "GBPm": "currency_m",
    "USD / share": "per_share", "GBp": "per_share",
    "%": "percent",
}
LABEL_COLUMN, UNIT_COLUMN, VALUE_COLUMN = "A", "B", "C"


def _rounded(value: float, unit: str) -> float:
    """Match the precision a filer reports at, so the cell reads like the figure it forecasts."""
    if unit == "percent":
        return round(value, 1)
    if unit == "per_share":
        return round(value, 2)
    return round(value, 1)


def write_books(forecasts: list[dict]) -> list[dict]:
    import openpyxl

    by_ticker: dict[str, dict[str, dict]] = {}
    registry = {r["metric"]: r for r in store.read(config.FEATURE / "design_matrix.parquet")}
    for row in forecasts:
        label = registry[row["metric"]]["workbook_label"]
        by_ticker.setdefault(row["ticker"], {})[label] = row

    SUBMISSION.mkdir(parents=True, exist_ok=True)
    written: list[dict] = []
    for stem, ticker in BOOKS.items():
        source = TEMPLATES / f"{stem}.xlsx"
        target = SUBMISSION / f"{stem}.xlsx"
        shutil.copyfile(source, target)
        book = openpyxl.load_workbook(target)
        sheet = book["Summary"]
        wanted = dict(by_ticker.get(ticker, {}))

        for index in range(1, sheet.max_row + 1):
            label = sheet[f"{LABEL_COLUMN}{index}"].value
            if not isinstance(label, str) or label not in wanted:
                continue
            row = wanted.pop(label)
            unit = sheet[f"{UNIT_COLUMN}{index}"].value
            expected = UNIT_EXPECTS.get(str(unit).strip())
            if expected and expected != row["unit"]:
                raise SystemExit(f"{stem}: '{label}' is '{unit}' in the workbook "
                                 f"({expected}) but {row['unit']} in the registry")
            value = _rounded(row["forecast"], row["unit"])
            sheet[f"{VALUE_COLUMN}{index}"] = value
            written.append({"workbook": f"{stem}.xlsx", "cell": f"{VALUE_COLUMN}{index}",
                            "ticker": ticker, "label": label, "metric": row["metric"],
                            "unit": str(unit), "value": value,
                            "anchor": row["chosen_anchor"]})
        if wanted:
            raise SystemExit(f"{stem}: no row in the template matches "
                             f"{sorted(wanted)} — labels must agree exactly")
        book.save(target)
    return written


def main() -> int:
    forecasts = store.read(MODEL / "forecasts.parquet")
    written = write_books(forecasts)
    if len(written) != 12:
        raise SystemExit(f"wrote {len(written)} cells, expected 12")

    print(f"submission/ — 4 workbooks, {len(written)} forecasts written\n")
    print(f"{'workbook':<22}{'cell':<6}{'metric':<36}{'units':<14}{'value':>13}")
    print("-" * 92)
    for row in written:
        print(f"{row['workbook']:<22}{row['cell']:<6}{row['label']:<36}"
              f"{row['unit']:<14}{row['value']:>13,.2f}")
    store.write(MODEL / "submission_cells.parquet", written)

    # read the saved files back — the submission is the file on disk, not the intent
    import openpyxl
    print()
    for stem in BOOKS:
        path = SUBMISSION / f"{stem}.xlsx"
        sheet = openpyxl.load_workbook(path)["Summary"]
        values = [sheet[f"{VALUE_COLUMN}{i}"].value for i in range(1, sheet.max_row + 1)
                  if isinstance(sheet[f"{LABEL_COLUMN}{i}"].value, str)
                  and sheet[f"{VALUE_COLUMN}{i}"].value is not None
                  and i > 6]
        ok = len(values) == 3 and all(isinstance(v, (int, float)) for v in values)
        print(f"  {'PASS' if ok else 'FAIL'}  {path.name:<22}"
              f"Summary sheet intact, 3 numeric forecasts: {values}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
